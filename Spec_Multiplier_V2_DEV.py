# This is a development test deployment.

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from io import BytesIO
import re
import time
import zipfile
import requests

st.set_page_config(page_title="Spec Constraint Controller", layout="wide")

# ---------------------------
# DARK UI STYLING
# ---------------------------
st.markdown("""
<style>
body { background-color: #0E1117; color: white; }

.stProgress > div > div > div > div {
    background-color: #00C6FF;
}

.green-time {
    color: #2ECC71;
    font-weight: 500;
    margin-top: 12px;
}
</style>
""", unsafe_allow_html=True)

st.title("Automated Batch Spec Constraint Controller V2.0")
st.info("Now supports different projects. You just need to specify the column where your contribution starts and where it ends.")
st.warning("Still in development Stage. Feel free to try it out if it works, if not god bless your soul")
st.markdown(
    """
    <div style='text-align: center; margin-top: 10px; margin-bottom: 10px;'>
        <a href="https://github.com/Rahul-2305/Automated_Spec_Constrainer/tree/main" target="_blank">
            <button style="
                background-color:#063970;
                color:white;
                padding:10px 20px;
                border:none;
                border-radius:8px;
                font-size:16px;
                cursor:pointer;">
                HOW TO USE ?
            </button>
        </a>
    </div>
    """,
    unsafe_allow_html=True
)

# ---------------------------
# FILE UPLOADS
# ---------------------------
uploaded_files = st.file_uploader(
    "Upload Spec Files",
    type=["xlsx"],
    accept_multiple_files=True
)

factor_file = st.file_uploader(
    "Upload Factor File (Multiple Sheets)",
    type=["xlsx"]
)

# ---------------------------
# SAMPLE TEMPLATE DOWNLOAD
# ---------------------------
@st.cache_data
def load_sample_template(url):
    response = requests.get(url)
    return response.content if response.status_code == 200 else None

sample_url = "https://github.com/Rahul-2305/Automated_Spec_Constrainer/raw/refs/heads/main/Factor_Sheet.xlsx"
sample_content = load_sample_template(sample_url)

if sample_content:
    st.download_button(
        label="⬇ Download Factor File Sample Template",
        data=sample_content,
        file_name="Factor_File_Sample_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------------------
# COLUMN RANGE SELECTOR
# ---------------------------
st.subheader("Select Contribution Column Range")

col1, col2 = st.columns(2)

with col1:
    start_column_letter = st.text_input(
        "Start Column (Excel Letter)",
        value="L"
    ).upper()

with col2:
    end_column_letter = st.text_input(
        "End Column (Excel Letter)",
        value="AC"
    ).upper()

# Validate Column Input
try:
    start_col = column_index_from_string(start_column_letter)
    end_col = column_index_from_string(end_column_letter)

    if start_col >= end_col:
        st.error("End column must be after Start column.")
        st.stop()

except:
    st.error("Invalid Excel column letter.")
    st.stop()

# ---------------------------
# PROCESSING
# ---------------------------
if uploaded_files and factor_file:

    factor_excel = pd.ExcelFile(factor_file)
    available_sheets = factor_excel.sheet_names

    st.subheader("Select Factor Sheet For Each Spec File")

    spec_sheet_map = {}

    header_col1, header_col2 = st.columns([3, 2])
    header_col1.markdown("**Spec File**")
    header_col2.markdown("**Factor Sheet**")

    st.markdown("---")

    for file in uploaded_files:

        row_col1, row_col2 = st.columns([3, 2])

        with row_col1:
            st.markdown(
                f"<div style='padding-top:8px; font-size:16px; font-weight:500;'>"
                f"{file.name}</div>",
                unsafe_allow_html=True
            )

        with row_col2:
            selected_sheet = st.selectbox(
                "",
                options=available_sheets,
                key=file.name
            )

        spec_sheet_map[file.name] = selected_sheet

    if st.button("Start Processing"):

        start_time = time.time()
        progress_bar = st.progress(0)
        progress_text = st.empty()

        total_steps = len(uploaded_files)
        current_step = 0

        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

            for spec_file in uploaded_files:

                selected_sheet = spec_sheet_map[spec_file.name]

                factor_df = pd.read_excel(
                    factor_file,
                    sheet_name=selected_sheet
                )
                factor_df.columns = factor_df.columns.str.strip()

                cols = list(factor_df.columns)
                left_cols = cols[:3]
                right_cols = cols[-2:]

                wb = load_workbook(spec_file)
                ws = wb["Base Hypothesis"]
                model_ws = wb["Model Specifications"]

                variable_col = 11
                year_row = 3
                data_start_row = 5

                # BASE HYPOTHESIS
                for _, row_data in factor_df.iterrows():

                    variable = str(row_data[left_cols[0]]).strip().lower() if pd.notna(row_data[left_cols[0]]) else None
                    factor_value = row_data[left_cols[1]] if pd.notna(row_data[left_cols[1]]) else None
                    constrainer = str(row_data[left_cols[2]]).strip() if pd.notna(row_data[left_cols[2]]) else None

                    if variable and constrainer:

                        for row in range(data_start_row, ws.max_row + 1):

                            cell_value = ws.cell(row=row, column=variable_col).value

                            if cell_value and str(cell_value).strip().lower() == variable:

                                if constrainer.upper() == "M":

                                    for col in range(start_col, end_col + 1):
                                        cell = ws.cell(row=row, column=col)
                                        if isinstance(cell.value, (int, float)):
                                            cell.value *= factor_value
                                            cell.number_format = "0.00%"

                                else:

                                    for col in range(start_col, end_col + 1):
                                        ws.cell(row=row, column=col).value = 0
                                        ws.cell(row=row, column=col).number_format = "0.00%"

                                    for col in range(start_col, end_col + 1):

                                        year_cell = ws.cell(row=year_row, column=col).value

                                        if year_cell and constrainer.lower() in str(year_cell).lower():

                                            ws.cell(row=row, column=col).value = factor_value
                                            ws.cell(row=row, column=col).number_format = "0.00%"

                                            if col + 1 <= end_col:
                                                ws.cell(row=row, column=col + 1).value = factor_value
                                                ws.cell(row=row, column=col + 1).number_format = "0.00%"

                                            break
                                break

                # MODEL SPECIFICATIONS
                for _, row_data in factor_df.iterrows():

                    model_variable = str(row_data[right_cols[0]]).strip().lower() if pd.notna(row_data[right_cols[0]]) else None
                    yesno_value = str(row_data[right_cols[1]]).strip() if pd.notna(row_data[right_cols[1]]) else None

                    if model_variable and yesno_value.lower() in ["yes", "no"]:

                        for row in range(1, model_ws.max_row + 1):

                            model_var = model_ws.cell(row=row, column=4).value

                            if model_var and str(model_var).strip().lower() == model_variable:

                                model_ws.cell(row=row, column=5).value = yesno_value.capitalize()
                                break

                # VERSION INCREMENT
                original_name = spec_file.name
                match = re.search(r'v(\d+)', original_name, re.IGNORECASE)

                if match:
                    version = int(match.group(1)) + 1
                    new_name = re.sub(r'v\d+', f'V{version}', original_name, flags=re.IGNORECASE)
                else:
                    new_name = original_name.replace(".xlsx", "_V2.xlsx")

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                zip_file.writestr(new_name, output.read())

                current_step += 1
                percent = int((current_step / total_steps) * 100)
                progress_bar.progress(percent)
                progress_text.markdown(f"**Progress: {percent}%**")

        end_time = time.time()
        elapsed_time = round(end_time - start_time, 2)

        progress_bar.progress(100)
        progress_text.markdown("**Progress: 100%**")

        st.markdown(
            f"<div class='green-time'>⏱ Execution Time: {elapsed_time} seconds</div>",
            unsafe_allow_html=True
        )

        st.markdown("<br>", unsafe_allow_html=True)

        zip_buffer.seek(0)

        st.download_button(
            label="⬇ Download All Processed Specs (ZIP)",
            data=zip_buffer,
            file_name="Processed_Spec_Files.zip",
            mime="application/zip"
        )

with st.expander("About this App"):
    st.write("Created by Beeraboina Rahul")
    st.write("Made in Python & Streamlit")
    st.write("Know more about Beeraboina Rahul at https://rahul-2305.github.io/Website/")

left, center, right = st.columns([1, 2, 1])

with center:
    col1, col2 = st.columns(2)

    with col1:
        if st.button("🎈 Want some Balloons"):
            st.balloons()

    with col2:
        if st.button("❄️ Want some Snow"):
            st.snow()

st.caption("© 2026 Beeraboina Rahul")
